import time
import openpyxl
import nltk
import re
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory

workbook = openpyxl.load_workbook('data/Data.xlsx')
stopwordtambahan=['jokowi','prabowo','pdip','kpu','golkar','ahok','paloh','sandiaga','amin',
                  'pilpres','kimi','hime','indonesia','salah','okezone','airlangga','hahahaha',
                  'subianto','partai','politik','salahkan','apasih','kalo','liputan','berita',
                  'detikcom','vivanews','pilkada','bawaslu','kpu','depok','radar','bandung','nasional',
                  'pdi-p','wali','kota','antaranews','negara','yusril','demokrat','yunus','istri','soeharto',
                  'titiek','bambang','presiden','pemilihan','megawati','nasdem','masuk','jakarta',
                  'tanggal','bahasa','nilai','risma','hadir','trilyun','ribu','selamat','wuryanto',
                  'selesai','iriana','enak','caleg','ganjar','vlog','anies','baswedan','pemimpin',
                  'islam','capres','orang','ketua','irashaimaseeeeeeeeeee','https','oleh',
                  'kali','jalan','joko','widodo','jend','refly','harun','hasil','juta','pemilu',
                  'merah','putus','situs','data','lengkap','resmi','sby-lbp-hp','pilih','kursi',
                  'jawa','dasar','atas','perintah','rakyat','menteri','sesiai','april','terima',
                  'tulis','terkait','kait','temu','duduk','jilid','beda','sadar','pdip-mega-bg',
                  'bodoh','betah','lembaga','bentuk','dorong','bentuk','anggota','suara','sesuai',
                  'bunga','tetap','bilang','dukung','muka','daya','besar','tempat','bingung','pasca',
                  'kembang','sulit','paham','dprd','jaring','khawatir','serah','guna','genk','kemarin',
                  'cermat','timbul','jahat','sambut','pimpin','guna','khusus','analis','batal','mudah',
                  'saji','buka','tuju','rawan','rumit','sistem','paksa','manipulasi','sistem','persen','umum',
                  'posisi','singgung','ribet','anggap','bubar','larang','selenggara','juang','udah',
                  'aceh','periode','nama','posisi'

                  ]

def IPreProcessing():
    print("apakah anda ingin melakukan Pre prosessing data?")
    AmbilDataYoN = input("(y/n/x):")
    if AmbilDataYoN == "y" or AmbilDataYoN == "Y":
        print("anda anda yakin ingin melakukan preprosesing? data yang lama akan terhapus?")
        AmbilDataYoN1 = input("(y/...):")
        if AmbilDataYoN1 == "y" or AmbilDataYoN1 =="Y":
            StartTime=time.time()
            PLowerCase()
            PTokenization()
            PStopWords()
            PStemming()
            EndTime=time.time()
            print("proses preprosesing keseluruhan telah selesai")
            print("waktu yang digunakan adalah :", EndTime - StartTime)
        else:
            return 1
    elif AmbilDataYoN == "n" or AmbilDataYoN == "N":
        print("anda tidak melakukan preprosesing")
    elif AmbilDataYoN == "x" or AmbilDataYoN == "X":
        return 0
    else:
        print("maaf input yang anda masukkan salah")
        IPreProcessing()


def replace_all(text, dic):
    for i, j in dic.items():
        text = text.replace(i, j)
    return text
def tokenfiltering(token):
    if len(token) < 4:
        token = token.replace(token, "")
    else:
        token=token

    token = re.sub(r'^https?:\/\/.*[\r\n]*', '', token, flags=re.MULTILINE)
    token = re.sub(r'^@.*[\r\n]*', '', token, flags=re.MULTILINE)
    token = re.sub(r'^#.*[\r\n]*', '', token, flags=re.MULTILINE)
    return token

def PLowerCase():
    # exel seting
    if 'PreLowerCase' in workbook.sheetnames:
        worksheet1=workbook['PreLowerCase']
    else:
        worksheet1 = workbook.create_sheet('PreLowerCase')

    worksheet1.cell(row=1, column=1).value = 'Data'
    worksheet1.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    worksheet1.cell(row=1, column=2).value = 'User'
    worksheet1.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    worksheet1.cell(row=1, column=3).value = 'Tweet'
    worksheet1.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    worksheet1.column_dimensions['A'].width = 30
    worksheet1.column_dimensions['B'].width = 30
    worksheet1.column_dimensions['C'].width = 150
    ExelRow = 2
    # txt seting
    PreProcessingText=  open("data/PreProcessingText.txt", "w+", encoding="utf-8")
    # proses lowercase dan penulisan
    StartTime=time.time()
    with open("data/raw.txt", "r", encoding="utf-8") as fileinput:
        for line in fileinput:
            line =line.rstrip()
            date, user, tweet = line.split('###',3)
            tweet=tweet.lower()
            PreProcessingText.write(str(date)+"###"+str(user)+"###"+str(tweet)+"\n")
            worksheet1.cell(row=ExelRow, column=1).value = date
            worksheet1.cell(row=ExelRow, column=2).value = user
            worksheet1.cell(row=ExelRow, column=3).value = tweet
            ExelRow = ExelRow + 1

    workbook.save('data/Data.xlsx')
    workbook.close()
    EndTime=time.time()
    PreProcessingText.close()
    print("proses PreProsesing-LowerCase selesai")
    print("waktu yang digunakan adalah :",EndTime-StartTime)

def PTokenization():
    # exel seting
    if 'PreTokenization' in workbook.sheetnames:
        worksheet2=workbook['PreTokenization']
    else:
        worksheet2 = workbook.create_sheet('PreTokenization')
    worksheet2.cell(row=1, column=1).value = 'Data'
    worksheet2.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    worksheet2.cell(row=1, column=2).value = 'User'
    worksheet2.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    worksheet2.cell(row=1, column=3).value = 'Tweet'
    worksheet2.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    worksheet2.column_dimensions['A'].width = 30
    worksheet2.column_dimensions['B'].width = 30
    worksheet2.column_dimensions['C'].width = 150
    ExelRow = 2
    # tokenization seting
    tkz=nltk.TweetTokenizer()
    # txt seting
    # UnUseTokenizationText = open("data/UnUseTokenizationText.txt", "w+", encoding="utf-8")
    TokenizationText = open("data/TokenizationText.txt", "w+", encoding="utf-8")
    # proses tokenisasi dan penulisan
    StartTime = time.time()
    with open("data/PreProcessingText.txt", "r", encoding="utf-8") as fileinput:
        for line in fileinput:
            line = line.rstrip()
            date, user, tweet = line.split('###', 3)
            tweet =tkz.tokenize(tweet)
            worksheet2.cell(row=ExelRow, column=1).value = date
            worksheet2.cell(row=ExelRow, column=2).value = user
            worksheet2.cell(row=ExelRow, column=3).value = str(tweet)
            TokenizationText.write(str(date)+"###"+str(user)+"###"+str(tweet)+"\n")
            # TokenizationText.write(str(tweet))
            ExelRow = ExelRow + 1

    workbook.save('data/Data.xlsx')
    workbook.close()
    EndTime = time.time()
    # UnUseTokenizationText.close()
    TokenizationText.close()
    print("proses PreProsesing-Tokenisasi selesai")
    print("waktu yang digunakan adalah :", EndTime - StartTime)
def PStopWords():
    replacedic = {"[": "", "]": "", ",": "", "'": ""}
    # exel seting
    if 'PreStopWords' in workbook.sheetnames:
        worksheet3=workbook['PreStopWords']
    else:
        worksheet3 = workbook.create_sheet('PreStopWords')
    worksheet3.cell(row=1, column=1).value = 'Data'
    worksheet3.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    worksheet3.cell(row=1, column=2).value = 'User'
    worksheet3.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    worksheet3.cell(row=1, column=3).value = 'Tweet'
    worksheet3.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    worksheet3.column_dimensions['A'].width = 30
    worksheet3.column_dimensions['B'].width = 30
    worksheet3.column_dimensions['C'].width = 150
    ExelRow = 2
    # Stopwordseting seting
    stopwords = nltk.corpus.stopwords.words('indonesian')
    # txt seting
    StopwordsText = open("data/StopwordsText.txt", "w+", encoding="utf-8")
    # proses stopwod dan penulisan
    StartTime = time.time()
    with open("data/TokenizationText.txt", "r", encoding="utf-8") as fileinput:
        for line in fileinput:
            dataoutput = []
            date, user, tweet = line.split('###', 3)
            data = replace_all(tweet, replacedic)
            tokens = [word for word in data.split()]
            # print(tokens)
            for token in tokens:
                token = tokenfiltering(token)
                if re.search('[a-zA-Z]', token) and (token not in stopwords) and (token not in stopwordtambahan):
                    dataoutput.append(token)
            worksheet3.cell(row=ExelRow, column=1).value = date
            # print(dataoutput)
            worksheet3.cell(row=ExelRow, column=2).value = user
            worksheet3.cell(row=ExelRow, column=3).value = str(dataoutput)
            StopwordsText.write(str(date) + "###" + str(user) + "###" + str(dataoutput) + "\n")
            ExelRow = ExelRow + 1

    workbook.save('data/Data.xlsx')
    workbook.close()
    EndTime = time.time()
    StopwordsText.close()
    print("proses PreProsesing-StopWords selesai")
    print("waktu yang digunakan adalah :", EndTime - StartTime)
def PStemming():
    replacedic = {"[": "", "]": "", ",": "", "'": ""}
    # exel seting
    if 'PreSteming' in workbook.sheetnames:
        worksheet = workbook['PreSteming']
    else:
        worksheet = workbook.create_sheet('PreSteming')
    worksheet.cell(row=1, column=1).value = 'Data'
    worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    worksheet.cell(row=1, column=2).value = 'User'
    worksheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    worksheet.cell(row=1, column=3).value = 'Tweet'
    worksheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 150
    ExelRow = 2
    # Stemming seting
    factory = StemmerFactory()
    stemmer = factory.create_stemmer()
    # txt seting
    PreStemingText = open("data/PreStemingText.txt", "w+", encoding="utf-8")
    # proses stemming dan penulisan
    StartTime = time.time()
    with open("data/StopwordsText.txt", "r", encoding="utf-8") as fileinput:
        for line in fileinput:
            dataoutput = []
            date, user, tweet = line.split('###', 3)
            # print(tweet)
            data = replace_all(tweet, replacedic)
            # print(data)
            hasil=stemmer.stem(data)
            # print(hasil)
            tokens = [word for word in hasil.split()]
            # print(tokens)
            for token in tokens:
                token=tokenfiltering(token)
                # dataoutput.append(token)
                if re.search('[a-zA-Z]', token) and (token not in stopwordtambahan):
                    dataoutput.append(token)
            # print(dataoutput)
            worksheet.cell(row=ExelRow, column=1).value = date
            # print(dataoutput)
            worksheet.cell(row=ExelRow, column=2).value = user
            worksheet.cell(row=ExelRow, column=3).value = str(dataoutput)
            PreStemingText.write(str(date) + "###" + str(user) + "###" + str(dataoutput) + "\n")
            ExelRow = ExelRow + 1

    workbook.save('data/Data.xlsx')
    workbook.close()
    EndTime = time.time()
    PreStemingText.close()
    print("proses PreProsesing-Steming selesai")
    print("waktu yang digunakan adalah :", EndTime - StartTime)

IPreProcessing()