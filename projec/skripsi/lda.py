import re
import time
import gensim
from gensim.models import CoherenceModel
import openpyxl
import pyLDAvis
import pyLDAvis.gensim  # don't skip this
import matplotlib.pyplot as plt
# matplotlib inline
from pprint import pprint
import importlib
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.manifold import MDS
import sklearn
from sklearn.decomposition import LatentDirichletAllocation, TruncatedSVD
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from multiprocessing import process,freeze_support
from sklearn.manifold import TSNE
from bokeh.plotting import figure, output_file, show
from bokeh.models import Label
from bokeh.io import output_notebook
import logging
from collections import Counter
import numpy as np
logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.ERROR)
from matplotlib.patches import Rectangle
import matplotlib.colors as mcolors

topicnumber = 10



import warnings
warnings.filterwarnings("ignore",category=DeprecationWarning)
workbook = openpyxl.load_workbook('data/Data.xlsx')
output_file('index.html')

AmbilDataYoN = 'n'
def ILdaProses():
    if __name__ == "__main__":
        freeze_support()
        print("anda menjalankan modul lda proses")
        PLdaProses()

# def ILdaProses():
#     print("anda menjalankan modul lda proses")
#     PLdaProses()

def PLdaProses():
    def replace_all(text, dic):
        for i, j in dic.items():
            text = text.replace(i, j)
        return text

    StartTime = time.time()
    datafordic = []
    replacedic = {"[": "", "]": "", ",": "", "'": "", "(": "", ")": ""}
    with open("data/PreStemingText.txt", "r", encoding="utf-8") as fileinput:
        for line in fileinput:
            dataoutput = []
            date, user, tweet = line.split('###', 3)
            # print(tweet)
            data = replace_all(tweet, replacedic)
            # print(data)
            tokens = [word for word in data.split()]
            # print(tokens)
            for token in tokens:
                dataoutput.append(token)
            # print(dataoutput)
            datafordic.append(dataoutput)

    # print(datafordic[:1])
    # print(datafordic)
    jumlahdoc=len(datafordic)
    # print(jumlah)
    # dictionery
    id2word = gensim.corpora.Dictionary(datafordic)
    # print(id2word)
    #menghilangkan term yang muncul kurang dari 2 dokumen dan lebih dati 0.9xjumlah dokumen
    # 100
    id2word.filter_extremes(no_below=3,no_above=0.9)
    # 1000
    # id2word.filter_extremes(no_below=20,no_above=0.9)
    # 6000
    # id2word.filter_extremes(no_below=120,no_above=0.9)
    # print(id2word)
    CorpusManusiaText = open("data/CorpusManusiaText.txt", "w+", encoding="utf-8")
    CorpusManusiaText.write(str(id2word))
    CorpusManusiaText.close()

    # corpus
    corpustext = datafordic
    # print(corpustext)

    # termdocument
    corpus = [id2word.doc2bow(text) for text in corpustext]
    print(corpus)
    CorpusComputerText = open("data/CorpusComputerText.txt", "w+", encoding="utf-8")
    CorpusComputerText.write(str(corpus))
    CorpusComputerText.close()

    # print(corpus[:1])
    # print([[(id2word[id], freq) for id, freq in cp] for cp in corpus])

    # lda model

    # Build LDA model
    # lda_model_auto = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=10, alpha='auto', per_word_topics=True)
    lda_model = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=int(topicnumber), alpha='auto', per_word_topics=True)
    # lda_model_itterasi = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word,iterations=10000, num_topics=10, alpha='auto', per_word_topics=True)
    # lda_model = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=10, random_state=100,
    #                                             update_every=1, chunksize=100, passes=10, alpha='auto',
    #                                             per_word_topics=True)
    # lda_model_alpha = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=10, alpha=0.1, per_word_topics=True)
    # lda_model_alpha1 = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=10, alpha=10, per_word_topics=True)
    # lda_model_beta = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=10, beta=0.1, per_word_topics=True)
    # lda_model_beta1 = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=10, alpha=10, per_word_topics=True)
    # lda_model_terbaik = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=10, alpha=0.00001 , iterations=10000,per_word_topics=True)

    # Print the Keyword in the 10 topics
    pprint(lda_model.print_topics(num_topics=int(topicnumber)))
    doc_lda = lda_model[corpus]
    ldamodelwrite=lda_model.print_topics(num_topics=int(topicnumber))
    # lda text
    LdaModelText = open("data/LdaModelText.txt", "w+", encoding="utf-8")
    LdaModelText.write(str(ldamodelwrite))
    LdaModelText.close()
    # lda exel
    # exel seting
    if 'HasilLdaModel' in workbook.sheetnames:
        worksheet1 = workbook['HasilLdaModel']
    else:
        worksheet1 = workbook.create_sheet('HasilLdaModel')

    worksheet1.cell(row=1, column=1).value = 'Topik ke-'
    worksheet1.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    worksheet1.cell(row=1, column=2).value = 'LDA-Model'
    worksheet1.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    worksheet1.column_dimensions['A'].width = 30
    worksheet1.column_dimensions['B'].width = 150
    ExelRow = 2
    topikeke = 0
    # print(ldamodelwrite)
    hasilldaexls=str(ldamodelwrite).split(")")
    for hasilldaexl in hasilldaexls:
        hasilldaexl = replace_all(hasilldaexl, replacedic)
        worksheet1.cell(row=ExelRow, column=1).value = topikeke
        worksheet1.cell(row=ExelRow, column=2).value = str(hasilldaexl)
        ExelRow = ExelRow + 1
        topikeke = topikeke +1

    # workbook.save('data/Data.xlsx')
    # workbook.close()
    # Compute Perplexity
    # print("lda model auto")
    # print('\nPerplexity: ', lda_model_auto.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # coherence_model_lda = CoherenceModel(model=lda_model_auto, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)
    # print("lda model iterasi")
    # print('\nPerplexity: ', lda_model_itterasi.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # coherence_model_lda = CoherenceModel(model=lda_model_itterasi, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)
    # print("lda model lda")
    # print("lda model alpha")
    # print('\nPerplexity: ', lda_model_alpha.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # coherence_model_lda = CoherenceModel(model=lda_model_alpha, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)
    # print("lda model alpha 1")
    # print('\nPerplexity: ', lda_model_alpha1.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # coherence_model_lda = CoherenceModel(model=lda_model_alpha1, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)
    # print("lda model beta")
    # print('\nPerplexity: ', lda_model_beta.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # coherence_model_lda = CoherenceModel(model=lda_model_beta, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)
    # print("lda model beta1")
    # print('\nPerplexity: ', lda_model_beta1.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # coherence_model_lda = CoherenceModel(model=lda_model_beta1, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)
    # print("lda model terbaik")
    # print('\nPerplexity: ', lda_model_terbaik.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # coherence_model_lda = CoherenceModel(model=lda_model_terbaik, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)

    # # real use lda
    # print('\nPerplexity: ', lda_model.log_perplexity(corpus))  # a measure of how good the model is. lower the better.
    # # Compute Coherence Score
    # coherence_model_lda = CoherenceModel(model=lda_model, texts=corpustext, dictionary=id2word, coherence='c_v')
    # coherence_lda = coherence_model_lda.get_coherence()
    # print('\nCoherence Score: ', coherence_lda)
    EndTime = time.time()
    print("proses pembuatan model lda selesai")
    print("waktu yang digunakan adalah :", EndTime - StartTime)
    # lda_model[corpus[11]]  # corpus[0] means the first document.
    # print(corpustext[0])
    # print("====")
    # print(corpus[0])
    # print(lda_model[corpus[1]])
    print("===============")
    print("apakah anda ingin melakukan penghitungan dengan jumlah topik yang optimal?")
    AmbilDataYoN = input("(y/n):")
    if AmbilDataYoN == "y" or AmbilDataYoN == "Y":
        print("optimalisasi jumlah topik sedang proses mohon menunggu...")
        StartTime1= time.time()
        # number of topic optimalisasi

        def compute_coherence_values(dictionary, corpus, texts, limit, start=2, step=3):
            """
            Compute c_v coherence for various number of topics

            Parameters:
            ----------
            dictionary : Gensim dictionary
            corpus : Gensim corpus
            texts : List of input texts
            limit : Max num of topics

            Returns:
            -------
            model_list : List of LDA topic models
            coherence_values : Coherence values corresponding to the LDA model with respective number of topics
            """
            coherence_values = []
            model_list = []
            for num_topics in range(start, limit, step):
                # model = gensim.models.wrappers.LdaMallet(mallet_path, corpus=corpus, num_topics=num_topics, id2word=id2word)
                model = gensim.models.ldamodel.LdaModel( corpus=corpus, num_topics=num_topics, id2word=id2word)
                model_list.append(model)
                coherencemodel = CoherenceModel(model=model, texts=texts, dictionary=dictionary, coherence='c_v')
                coherence_values.append(coherencemodel.get_coherence())

            return model_list, coherence_values

        # Can take a long time to run.
        # 100
        OptimilisasiStart = 2
        OptimilisasiLimit = 10
        OptimilisasiStep = 1
        # # 1000
        # OptimilisasiStart = 2
        # OptimilisasiLimit = 100
        # OptimilisasiStep = 10
        # 6000
        # OptimilisasiStart= 2
        # OptimilisasiLimit= 600
        # OptimilisasiStep= 60

        model_list, coherence_values = compute_coherence_values(dictionary=id2word, corpus=corpus, texts=corpustext,
                                                                start=OptimilisasiStart, limit=OptimilisasiLimit, step=OptimilisasiStep)
        # Show graph
        limit = OptimilisasiLimit;
        start = OptimilisasiStart;
        step = OptimilisasiStep;
        x = range(start, limit, step)
        plt.plot(x, coherence_values)
        plt.xlabel("Num Topics")
        plt.ylabel("Coherence score")
        plt.legend(("coherence_values"), loc='best')
        plt.show()

        # Print the coherence scores
        OptimalTopic= 0
        OptimalTopicArray= 0
        OptimalCounter=0
        OptimalTopicNumber=0
        for m, cv in zip(x, coherence_values):
            print("Num Topics =", m, " has Coherence Value of", round(cv, 4))
            OptimalTopicTemp=round(cv,4)
            if OptimalTopicTemp>OptimalTopic:
                OptimalTopic=OptimalTopicTemp
                OptimalTopicArray=OptimalCounter
                OptimalTopicNumber=m
            else:
                OptimalTopic=OptimalTopic
                OptimalTopicArray=OptimalTopicArray
                OptimalTopicNumber=OptimalTopicNumber
            OptimalCounter = OptimalCounter+1
        print(OptimalTopicArray)
        print(OptimalTopic)
        print(OptimalTopicNumber)
        # Select the model and print the topics
        # optimal_model = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=int(OptimalTopicNumber), iterations=10000, random_state=100,
        #                                             update_every=1, chunksize=100, passes=10, alpha='auto',
        #                                             per_word_topics=True)
        optimal_model = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=int(OptimalTopicNumber),
                                                    alpha='auto', per_word_topics=True)
        # optimal_model = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, iterations=10000, num_topics=int(OptimalTopicNumber),
        #                                             alpha='auto', per_word_topics=True)

        # model_topics = optimal_model.show_topics(formatted=False)
        pprint(optimal_model.print_topics(num_topics=int(OptimalTopicNumber)))
        ldamodeloptimalwrite = optimal_model.print_topics(num_topics=int(OptimalTopicNumber))

        # lda text
        LdaModelOptimalText = open("data/LdaModelOptimalText.txt", "w+", encoding="utf-8")
        LdaModelOptimalText.write(str(ldamodeloptimalwrite))
        LdaModelOptimalText.close()
        # lda exel
        # exel seting
        if 'HasilLdaModelOptimal' in workbook.sheetnames:
            HasilLdaModelOptimal = workbook['HasilLdaModelOptimal']
        else:
            HasilLdaModelOptimal = workbook.create_sheet('HasilLdaModelOptimal')

        HasilLdaModelOptimal.cell(row=1, column=1).value = 'Topik ke-'
        HasilLdaModelOptimal.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        HasilLdaModelOptimal.cell(row=1, column=2).value = 'LDA-Model'
        HasilLdaModelOptimal.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
        HasilLdaModelOptimal.column_dimensions['A'].width = 30
        HasilLdaModelOptimal.column_dimensions['B'].width = 150
        ExelRow = 2
        topikeke = 0
        # print(ldamodelwrite)
        hasilldaexlsoptimal = str(ldamodeloptimalwrite).split(")")
        for hasilldaexl1 in hasilldaexlsoptimal:
            hasilldaexl1 = replace_all(hasilldaexl1, replacedic)
            HasilLdaModelOptimal.cell(row=ExelRow, column=1).value = topikeke
            HasilLdaModelOptimal.cell(row=ExelRow, column=2).value = str(hasilldaexl1)
            ExelRow = ExelRow + 1
            topikeke = topikeke + 1

        EndTime1 = time.time()
        print("proses optimalisasi jumlah topik telah selesai")
        print("waktu yang digunakan adalah :", EndTime1 - StartTime1)
        print("================================================================================")
    else:
        print("tidak melakukan optimalisasi")

    # visualisasi
    print("apakah anda ingin membuat diagram dari data?")
    AmbilDataYoN2 = input("(y/n):")
    if AmbilDataYoN2 == "y" or AmbilDataYoN2 == "Y":
        if AmbilDataYoN =="y" or AmbilDataYoN == "Y":
            print("")
            modell = optimal_model
            jumlahtopik=OptimalTopicNumber
        else:
            modell = lda_model
            jumlahtopik=topicnumber
        # Visualize the topics
        print("sedang melakukan proses visualisasi")
        StartTime3=time.time()
        print("lakukan diagram distribusi topik?")
        distribusitopikYoN=input("(y/n")
        if distribusitopikYoN == "y" or distribusitopikYoN == "Y":
            # Get topic weights
            topic_weights = []
            for i, row_list in enumerate(modell[corpus]):
                topic_weights.append([w for i, w in row_list[0]])

            # Array of topic weights
            arr = pd.DataFrame(topic_weights).fillna(0).values

            # Keep the well separated points (optional)
            arr = arr[np.amax(arr, axis=1) > 0.35]

            # Dominant topic number in each doc
            topic_num = np.argmax(arr, axis=1)

            # tSNE Dimension Reduction
            tsne_model = TSNE(n_components=2, verbose=1, random_state=0, angle=.99, init='pca')
            tsne_lda = tsne_model.fit_transform(arr)

            # Plot the Topic Clusters using Bokeh
            output_notebook()
            n_topics = int(jumlahtopik)
            mycolors = np.array([color for name, color in mcolors.TABLEAU_COLORS.items()])
            plot = figure(title="t-SNE Clustering of {} LDA Topics".format(n_topics),
                          plot_width=900, plot_height=700)
            plot.scatter(x=tsne_lda[:, 0], y=tsne_lda[:, 1], color=mycolors[topic_num])
            # plot.scatter(x=tsne_lda[:, 0], y=tsne_lda[:, 1], color=mycolors[jumlahtopik])
            show(plot)

            # chart
            topics = modell.show_topics(formatted=False)
            data_flat = [w for w_list in datafordic for w in w_list]
            counter = Counter(data_flat)

            out = []
            for i, topic in topics:
                for word, weight in topic:
                    out.append([word, i, weight, counter[word]])

        else:
            print("")
        # df = pd.DataFrame(out, columns=['word', 'topic_id', 'importance', 'word_count'])

        # # Plot Word Count and Weights of Topic Keywords
        # fig, axes = plt.subplots(2, 5, figsize=(50, 50), sharey=True, dpi=160)
        # cols = [color for name, color in mcolors.TABLEAU_COLORS.items()]
        # for i, ax in enumerate(axes.flatten()):
        #     ax.bar(x='word', height="word_count", data=df.loc[df.topic_id == i, :], color=cols[i], width=0.5, alpha=0.3,
        #            label='Word Count')
        #     ax_twin = ax.twinx()
        #     ax_twin.bar(x='word', height="importance", data=df.loc[df.topic_id == i, :], color=cols[i], width=0.2,
        #                 label='Weights')
        #     ax.set_ylabel('Word Count', color=cols[i])
        #     ax_twin.set_ylim(0, 1);
        #     ax.set_ylim(0, 100)
        #     ax.set_title('Topic: ' + str(i), color=cols[i], fontsize=16)
        #     ax.tick_params(axis='y', left=False)
        #     ax.set_xticklabels(df.loc[df.topic_id == i, 'word'], rotation=30, horizontalalignment='right')
        #     ax.legend(loc='upper left');
        #     ax_twin.legend(loc='upper right')
        #
        # fig.tight_layout(w_pad=5)
        # fig.suptitle('Word Count and Importance of Topic Keywords', fontsize=22, y=1.05)
        # plt.show()


        # pyLDAvis.enable_notebook()
        # vis = pyLDAvis.gensim.prepare(lda_model, corpus, dictionary=lda_model.id2word)
        # pyLDAvis.save_html(vis, 'lda.html')
        # vis

        # with open("data/LdaModelText.txt",'r', encoding="utf-8") as datainput:
        #     # str(datainput).replace(')','#')
        #     for line in datainput:
        #         print(line)
        #         datafordiagram=line.replace('), ','\n')
        #         print(datafordiagram)
        #         dataprosesfordaig=replace_all(datafordiagram,replacedic)
        #         print(dataprosesfordaig)
        topics_matrix = modell.show_topics(formatted=False)
        def plot_word_dist_per_topic(topic_no, topics_matrix, width=0.5):
            _, topic_words = topics_matrix[topic_no]

            # hanya top-10 yang paling tinggi probabilitasnya
            words = []
            probs = []
            for word, prob in topic_words:
                words.append(word)
                probs.append(prob)

            ind = np.arange(len(words))

            plt.bar(ind, probs, width=width)
            plt.xticks(ind + width / 2, words, rotation='vertical')
            plt.title('Word Distribution of Topic {}'.format(topic_no))
            plt.show()
        jumlahdiagram=0
        while jumlahdiagram<jumlahtopik:
            plot_word_dist_per_topic(jumlahdiagram, topics_matrix)
            jumlahdiagram=jumlahdiagram+1
        EndTime3=time.time()
        print("proses visualisasi tela selesai")
        print("waktu yang digunakan adalah :", EndTime3 - StartTime3)
    else:
        print("visualisasi tidak dilakukan")

    # evaluasi
    print("lakukan evaluasi?")
    AmbilDataYoN1 = input("(y/n):")
    if AmbilDataYoN1 =="y" or AmbilDataYoN1 =="Y":
        print("proses penulisan hasil sedang berjalan")
        StartTime2=time.time()
        print("====")
        hasilwritercounter = 0
        if AmbilDataYoN =="y" or AmbilDataYoN == "Y":
            print("")
            modell = optimal_model
        else:
            modell = lda_model

        if 'PersentaseTopik' in workbook.sheetnames:
            PersentaseTopik = workbook['PersentaseTopik']
        else:
            PersentaseTopik = workbook.create_sheet('PersentaseTopik')

        PersentaseTopik.cell(row=1, column=1).value = 'Date'
        PersentaseTopik.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        PersentaseTopik.cell(row=1, column=2).value = 'User'
        PersentaseTopik.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
        PersentaseTopik.cell(row=1, column=3).value = 'Tweet'
        PersentaseTopik.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
        PersentaseTopik.cell(row=1, column=4).value = 'Persentase'
        PersentaseTopik.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)
        PersentaseTopik.column_dimensions['A'].width = 30
        PersentaseTopik.column_dimensions['B'].width = 30
        PersentaseTopik.column_dimensions['C'].width = 150
        PersentaseTopik.column_dimensions['D'].width = 100

        if 'HasilOutput' in workbook.sheetnames:
            HasilOutput = workbook['HasilOutput']
        else:
            HasilOutput = workbook.create_sheet('HasilOutput')

        HasilOutput.cell(row=1, column=1).value = 'Date'
        HasilOutput.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        HasilOutput.cell(row=1, column=2).value = 'User'
        HasilOutput.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
        HasilOutput.cell(row=1, column=3).value = 'Tweet'
        HasilOutput.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
        HasilOutput.cell(row=1, column=4).value = 'Main topik'
        HasilOutput.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)
        HasilOutput.cell(row=1, column=5).value = 'Keyword'
        HasilOutput.cell(row=1, column=5).font = openpyxl.styles.Font(bold=True)
        HasilOutput.column_dimensions['A'].width = 30
        HasilOutput.column_dimensions['B'].width = 30
        HasilOutput.column_dimensions['C'].width = 150
        HasilOutput.column_dimensions['D'].width = 10
        HasilOutput.column_dimensions['E'].width = 100

        ExelRow = 2
        topikeke = 0
        HasilOutputText = open("data/HasilOutputText.txt", "w+", encoding="utf-8")
        PersentaseTopikText = open("data/PersentaseTopikText.txt", "w+", encoding="utf-8")
        with open("data/raw.txt", "r", encoding="utf-8") as fileinput:
            for line in fileinput:
                line = line.rstrip()
                date, user, tweet = line.split('###', 3)
                # print("dokumen ke", hasilwritercounter)
                # print topik persentase untuk
                # print(modell.get_document_topics(corpus[hasilwritercounter]))
                persentase=modell.get_document_topics(corpus[hasilwritercounter])
                # print(lda_model.get_topic_terms(0))
                # print(lda_model.get_term_topics(0))
                # print(lda_model.get_topics())
                # main topik
                maintopik = modell.get_document_topics(corpus[hasilwritercounter])
                # print(type(maintopik))
                # print(maintopik)
                # replacedic1 = {" ": "|", "]": "", "[": ""}
                # scoring = replace_all(str(maintopik), replacedic1)
                # print(scoring)
                topikutama = 0
                scorutama = 0
                for maintopiks in maintopik:
                    # print(maintopiks)
                    scoring = replace_all(str(maintopiks), replacedic)
                    # print(scoring)
                    topikke, scor = scoring.split(" ", 2)
                    # print(topikke)
                    # print(scor)
                    topikutamatemp = topikke
                    scoreutamatemp = scor
                    if float(scorutama) < float(scoreutamatemp):
                        topikutama = topikutamatemp
                        scorutama = scoreutamatemp
                    else:
                        topikutama = topikutama
                        scorutama = scorutama
                # print("topik utama adalah", topikutama, scorutama)
                if float(scorutama)<0.2:
                    topikutama=0-1
                    topik_keyword=['null']
                else:
                    topikutama=topikutama
                    wp = modell.show_topic(int(topikutama))
                    topik_keyword = ", ".join([word for word, prop in wp])


                # print(topik_keyword)
                print("==================")
                print("doc",hasilwritercounter)
                print("persentase")
                print(str(date) + "###" + str(user) + "###" + str(tweet) + "###" + str(persentase))
                print("hasil output")
                print(str(date) + "###" + str(user) + "###" + str(tweet) + "###" + str(topikutama) + "###" + str(topik_keyword))
                PersentaseTopikText.write(str(date) + "###" + str(user) + "###" + str(tweet) + "###" + str(persentase) + "\n")
                HasilOutputText.write(str(date) + "###" + str(user) + "###" + str(tweet) + "###" + str(topikutama) + "###" + str(topik_keyword) + "\n")
                PersentaseTopik.cell(row=ExelRow, column=1).value = date
                PersentaseTopik.cell(row=ExelRow, column=2).value = user
                PersentaseTopik.cell(row=ExelRow, column=3).value = tweet
                PersentaseTopik.cell(row=ExelRow, column=4).value = str(persentase)
                HasilOutput.cell(row=ExelRow, column=1).value = date
                HasilOutput.cell(row=ExelRow, column=2).value = user
                HasilOutput.cell(row=ExelRow, column=3).value = tweet
                HasilOutput.cell(row=ExelRow, column=4).value = str(topikutama)
                HasilOutput.cell(row=ExelRow, column=5).value = str(topik_keyword)
                ExelRow = ExelRow + 1
                hasilwritercounter = hasilwritercounter + 1
        EndTime2=time.time()
        print("proses penulisan hasil")
        print("waktu yang digunakan adalah :", EndTime2 - StartTime2)
        print("================================================================================")


    else:
        print("tidak melakukan evaluasi")
    workbook.save('data/Data.xlsx')
    workbook.close()
ILdaProses()








