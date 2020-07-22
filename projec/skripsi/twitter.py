import time
import tweepy
import openpyxl
import twitter_credentials
# import GetOldTweets3 as got
from TwitterAPI import TwitterAPI
# auth for twitter account
auth = tweepy.OAuthHandler(twitter_credentials.CONSUMER_KEY, twitter_credentials.CONSUMER_SECRET)
auth.set_access_token(twitter_credentials.ACCESS_TOKEN, twitter_credentials.ACCESS_TOKEN_SECRET)
# api = TwitterAPI(twitter_credentials.CONSUMER_KEY, twitter_credentials.CONSUMER_SECRET ,twitter_credentials.ACCESS_TOKEN, twitter_credentials.ACCESS_TOKEN_SECRET)
api = tweepy.API(auth,wait_on_rate_limit=True)
# api = tweepy.API(auth)


# seting xlsx file
workbook = openpyxl.Workbook()
# worksheet = workbook.create_sheet('RawData')


def IAmbilData():
    print("apakah anda ingin melakukan pengambilan data?")
    AmbilDataYoN= input("(y/n/x):")
    if AmbilDataYoN == "y" or AmbilDataYoN == "Y":
        print("anda anda yakin ingin mengambil data? data yang lama akan terhapus?")
        AmbilDataYoN1 = input("(y/...):")
        if AmbilDataYoN1 == "y" or AmbilDataYoN1 =="Y":
            StartTime=time.time()
            PAmbilData()
            EndTime=time.time()
            print("proses pengambilan data selesai")
            print("waktu yang digunakan adalah :", EndTime - StartTime)
        else:
            return 1
    elif AmbilDataYoN == "n" or AmbilDataYoN == "N":
        print("anda tidak ingin mengambil data")
    elif AmbilDataYoN == "x" or AmbilDataYoN == "X":
        return 0
    else:
        print("maaf input yang anda masukkan salah")
        IAmbilData()
def PAmbilData():
    # exel formated
    if 'RawData' in workbook.sheetnames:
        worksheet=workbook['RawData']
    else:
        worksheet = workbook.create_sheet('RawData')

    worksheet.cell(row=1,column=1).value='Data'
    worksheet.cell(row=1, column=1).font=openpyxl.styles.Font(bold=True)
    worksheet.cell(row=1, column=2).value = 'User'
    worksheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    worksheet.cell(row=1, column=3).value = 'Tweet'
    worksheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 150
    ExelRow = 2
    # txt file seting
    raw = open("data/raw.txt", "w+", encoding="utf-8")
    # twitter ambil data seting
    query = 'ibu kota baru'
    # query = 'pemindahan ibu kota indonesia'
    # query = '#ibukotabaru'
    # query = '#pemindahanibukota'
    # query = 'ibu kota baru indonesia'
    max_tweets = 2000
    # max_tweets = 1

    # proses pengambilan data
    searched_tweets = [status for status in tweepy.Cursor(api.search, q=query, tweet_mode='extended', lang="id", ).items(max_tweets)]

    for tweet in searched_tweets:
        print(tweet.created_at, "##", tweet.user.screen_name, "##", tweet.full_text, "\n")
        tweetenterremovel=tweet.full_text
        tweetenterremovel=tweetenterremovel.replace('\n', ' ')
        raw.write(str(tweet.created_at) + "###" + str(tweet.user.screen_name) + "###" + str(tweetenterremovel) + "\n")
        worksheet.cell(row=ExelRow,column=1).value=tweet.created_at
        worksheet.cell(row=ExelRow,column=2).value=tweet.user.screen_name
        worksheet.cell(row=ExelRow,column=3).value=tweet.full_text
        ExelRow = ExelRow + 1
        # print(dir(tweet))
        # print(tweet.created_at)
        # print(tweet.user)
    workbook.save('data/Data.xlsx')
    workbook.close()


    # AmbilJumlahData=10000
    # query = '#pemilu2019 OR #pilpres2019 OR #pileg2019'
    # tweetCriteria = got.manager.TweetCriteria().setQuerySearch(query) \
    #     .setSince("2019-01-01") \
    #     .setUntil("2019-07-30") \
    #     .setMaxTweets(int(AmbilJumlahData))
    # counter = 0
    # while counter<int(AmbilJumlahData):
    #     tweet = got.manager.TweetManager.getTweets(tweetCriteria)[counter]
    #     print(tweet.date ,tweet.username, tweet.text)
    #     # print(tweet.created_at, "##", tweet.user.screen_name, "##", tweet.full_text, "\n")
    #     # tweetenterremovel=tweet.full_text
    #     # tweetenterremovel=tweetenterremovel.replace('\n', ' ')
    #     raw.write(str(tweet.date) + "###" + str(tweet.username) + "###" + str(tweet.text) + "\n")
    #     worksheet.cell(row=ExelRow,column=1).value=tweet.date
    #     worksheet.cell(row=ExelRow,column=2).value=tweet.username
    #     worksheet.cell(row=ExelRow,column=3).value=tweet.text
    #     ExelRow = ExelRow + 1
    #     # print(dir(tweet))
    #     # print(tweet.created_at)
    #     # print(tweet.user)
    #
    #     counter=counter+1

    # query = '#pemilu2019 OR #pilpres2019 OR #pileg2019'
    # jumlahdatatweet=10
    # r = api.request('search/tweets', {'q': query, 'count':jumlahdatatweet})
    # for item in r.get_iterator():
    #     if 'text' in item:
    #         print(item['text'])
    #         print(item['created_at'])
    #         # print(item['screen_name'])
    #         print(dir(item))
    #         print(item)
    #         # print (item['create_at']+ item['name']+ item['text'])
    #         # raw.write(str(item['create_at']) + "###" + str(item['name']) + "###" + str(item['text']) + "\n")
    #         # worksheet.cell(row=ExelRow,column=1).value=str(item['create_at'])
    #         # worksheet.cell(row=ExelRow,column=2).value=str(item['name'])
    #         # worksheet.cell(row=ExelRow,column=3).value=str(item['text'])
    #         ExelRow = ExelRow + 1
    #     elif 'message' in item and item['code'] == 88:
    #         print('SUSPEND, RATE LIMIT EXCEEDED: %s\n' % item['message'])
    #         break
    #
    # workbook.close()
    raw.close()
    print("proses selesai")

IAmbilData()





